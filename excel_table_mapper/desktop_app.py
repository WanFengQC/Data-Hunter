from __future__ import annotations

import json
import os
import queue
import random
import shutil
import socket
import string
import sys
import threading
import time
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import httpx
import psycopg
from openpyxl import load_workbook

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except Exception:
    DND_FILES = None
    TkinterDnD = None

ROOT = Path(__file__).resolve().parent
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from excel_table_mapper.processor import (  # noqa: E402
    DEFAULT_STOPWORDS,
    build_table2_with_options,
    discover_phrase_compact_normalization_candidates,
    discover_plural_normalization_candidates,
    normalize_input_rows_with_map,
    read_input_rows,
    write_output_with_options,
)
from excel_table_mapper.tagger import PgConfig, Tagger  # noqa: E402


APP_TITLE = "TrafficAnalysisTool"
APP_DIR_NAME = "TrafficAnalysisTool"
LOCAL_CACHE_NAME = "traffic_analysis_tool_cache.sqlite3"
SEED_CACHE_NAME = "seed_word_cache.sqlite3"
DEFAULT_DB_HOST = "192.168.110.107"
DEFAULT_DB_PORT = 5432
DEFAULT_DB_NAME = "hunter"
DEFAULT_DB_USER = "postgres"
DEFAULT_DB_PASS = "123456"
DEFAULT_DB_SCHEMA = "public"
DEFAULT_DB_TABLE = "traffic_analysis_word_cache"
DEFAULT_AI_MODEL = "gpt-5.4"
DEFAULT_AI_BASE_URL = "https://api.uniapi.io"
DEFAULT_OPENAI_API_KEY = "sk-f93c5cc7df984f2ea501017091e1e633"
DEFAULT_BATCH_SIZE = 50


def _runtime_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return ROOT


def _resource_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(getattr(sys, "_MEIPASS", _runtime_base_dir()))
    return ROOT


def _app_data_dir() -> Path:
    base = _runtime_base_dir()
    path = base / APP_DIR_NAME
    path.mkdir(parents=True, exist_ok=True)
    return path


def _default_local_cache_db() -> Path:
    cache_dir = _app_data_dir() / "cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir / LOCAL_CACHE_NAME


def _normalize_cache_db_name(value: str | None) -> str:
    name = Path(str(value or "").strip() or LOCAL_CACHE_NAME).name
    if "." not in name:
        name = f"{name}.sqlite3"
    return name


def _resolve_local_cache_db_path(value: str | None) -> Path:
    name = _normalize_cache_db_name(value)
    cache_dir = _app_data_dir() / "cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir / name


def _default_output_for_input(input_path: Path | None) -> Path:
    output_dir = _runtime_base_dir() / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = (input_path.stem if input_path else "result").strip() or "result"
    rand = "".join(random.choices(string.ascii_lowercase + string.digits, k=6))
    return output_dir / f"{stem}_{rand}_output.xlsx"


def _seed_cache_path() -> Path:
    return _resource_dir() / "assets" / SEED_CACHE_NAME


def _ensure_local_cache_seed(local_db: Path) -> None:
    local_db.parent.mkdir(parents=True, exist_ok=True)
    if local_db.exists():
        return
    seed = _seed_cache_path()
    if seed.exists():
        shutil.copy2(seed, local_db)


def _can_connect(host: str, port: int, timeout_sec: float = 2.8, retries: int = 3, delay_sec: float = 0.35) -> bool:
    for i in range(max(1, retries)):
        try:
            with socket.create_connection((host, int(port)), timeout=timeout_sec):
                return True
        except Exception:
            if i < retries - 1:
                time.sleep(max(0.0, delay_sec))
    return False


def _load_json(path: Path, default: dict) -> dict:
    if not path.exists():
        return dict(default)
    try:
        return {**default, **json.loads(path.read_text(encoding="utf-8"))}
    except Exception:
        return dict(default)


def _save_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _is_masked_secret(value: str) -> bool:
    stripped = str(value or "").strip()
    return bool(stripped) and set(stripped) == {"*"}


def _extract_dropped_files(widget: tk.Misc, data: str) -> list[str]:
    text = str(data or "").strip()
    if not text:
        return []
    try:
        parts = list(widget.tk.splitlist(text))
    except Exception:
        parts = [text]
    out: list[str] = []
    for p in parts:
        val = str(p).strip()
        if val.startswith("{") and val.endswith("}"):
            val = val[1:-1]
        if val:
            out.append(val)
    return out


class App:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1200x760")

        self.queue: queue.Queue[tuple[str, dict]] = queue.Queue()
        self.is_processing = False
        self.sync_in_progress = False
        self.sync_done = False
        self.db_available = False
        self._suppress_mode_event = False

        self.config_path = _app_data_dir() / "settings.json"
        self.local_cache_default = _default_local_cache_db()
        _ensure_local_cache_seed(self.local_cache_default)

        defaults = {
            "mode": "offline",
            "enable_ai": True,
            "use_history_cache": True,
            "ai_model": DEFAULT_AI_MODEL,
            "ai_base_url": DEFAULT_AI_BASE_URL,
            "ai_batch_size": DEFAULT_BATCH_SIZE,
            "openai_api_key": DEFAULT_OPENAI_API_KEY,
            "local_cache_db": self.local_cache_default.name,
            "db_host": DEFAULT_DB_HOST,
            "db_port": DEFAULT_DB_PORT,
            "db_name": DEFAULT_DB_NAME,
            "db_user": DEFAULT_DB_USER,
            "db_pass": DEFAULT_DB_PASS,
            "db_schema": DEFAULT_DB_SCHEMA,
            "db_table": DEFAULT_DB_TABLE,
            "debug_mode": False,
            "enable_normalization": True,
            "enable_phrase_normalization": True,
            "enable_plural_normalization": True,
        }
        self.settings = _load_json(self.config_path, defaults)
        old_norm = bool(self.settings.get("enable_normalization", True))
        if "enable_phrase_normalization" not in self.settings:
            self.settings["enable_phrase_normalization"] = old_norm
        if "enable_plural_normalization" not in self.settings:
            self.settings["enable_plural_normalization"] = old_norm
        self.settings["local_cache_db"] = _normalize_cache_db_name(self.settings.get("local_cache_db"))
        if str(self.settings.get("db_table", "")).strip() == "seller_sprite_word_cache":
            self.settings["db_table"] = DEFAULT_DB_TABLE
        _save_json(self.config_path, self.settings)
        self.output_user_customized = False
        self._api_key_secret = "" if _is_masked_secret(str(self.settings.get("openai_api_key", ""))) else str(
            self.settings.get("openai_api_key", "") or ""
        )
        self._db_pass_secret = "" if _is_masked_secret(str(self.settings.get("db_pass", ""))) else str(
            self.settings.get("db_pass", "") or ""
        )
        self.last_generated_output_path: Path | None = None

        self.input_var = tk.StringVar(value="")
        self.sheet_var = tk.StringVar(value="")
        self.output_var = tk.StringVar(value=str(_default_output_for_input(None)))

        self.mode_var = tk.StringVar(value=self.settings.get("mode", "offline"))
        self.use_history_cache_var = tk.BooleanVar(value=bool(self.settings.get("use_history_cache", True)))
        self.debug_mode_var = tk.BooleanVar(value=bool(self.settings.get("debug_mode", False)))
        self.enable_phrase_normalization_var = tk.BooleanVar(
            value=bool(self.settings.get("enable_phrase_normalization", True))
        )
        self.enable_plural_normalization_var = tk.BooleanVar(
            value=bool(self.settings.get("enable_plural_normalization", True))
        )

        self.status_var = tk.StringVar(value="待处理")
        self.progress_text_var = tk.StringVar(value="")
        self.mode_desc_var = tk.StringVar(value="自动切换（可连数据库=云端，失败=本地）")
        self.last_db_failure_reason = ""

        self._build_ui()
        self.root.after(80, self._startup_probe_and_set_mode)
        self.root.after(100, self._poll_queue)

    def _build_ui(self) -> None:
        frm = ttk.Frame(self.root, padding=10)
        frm.pack(fill=tk.BOTH, expand=True)

        row = 0
        ttk.Label(frm, text="输入文件").grid(row=row, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(frm, textvariable=self.input_var).grid(row=row, column=1, sticky="ew", pady=4)
        ttk.Button(frm, text="选择", command=self._choose_input).grid(row=row, column=2, padx=(8, 0), pady=4)

        row += 1
        ttk.Label(frm, text="输入Sheet").grid(row=row, column=0, sticky="w", padx=(0, 8), pady=4)
        self.sheet_combo = ttk.Combobox(frm, textvariable=self.sheet_var, values=[], state="readonly")
        self.sheet_combo.grid(row=row, column=1, sticky="ew", pady=4)

        row += 1
        ttk.Label(frm, text="输出文件").grid(row=row, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(frm, textvariable=self.output_var).grid(row=row, column=1, sticky="ew", pady=4)
        ttk.Button(frm, text="另存为", command=self._choose_output).grid(row=row, column=2, padx=(8, 0), pady=4)

        row += 1
        mode_bar = ttk.Frame(frm)
        mode_bar.grid(row=row, column=0, columnspan=3, sticky="ew", pady=(8, 4))
        ttk.Label(mode_bar, text="运行方式：").pack(side=tk.LEFT)
        ttk.Label(mode_bar, textvariable=self.mode_desc_var).pack(side=tk.LEFT)
        ttk.Button(mode_bar, text="配置", command=self._open_settings_dialog).pack(side=tk.RIGHT)
        ttk.Button(mode_bar, text="归一化管理", command=self._open_norm_manager).pack(side=tk.RIGHT, padx=(0, 8))

        row += 1
        ai_bar = ttk.Frame(frm)
        ai_bar.grid(row=row, column=0, columnspan=3, sticky="ew", pady=4)
        self.history_cache_check = ttk.Checkbutton(ai_bar, text="启用历史缓存", variable=self.use_history_cache_var)
        self.history_cache_check.pack(side=tk.LEFT)
        ttk.Checkbutton(ai_bar, text="启用词组归一化", variable=self.enable_phrase_normalization_var).pack(
            side=tk.LEFT, padx=(10, 0)
        )
        ttk.Checkbutton(ai_bar, text="启用复数归一化", variable=self.enable_plural_normalization_var).pack(
            side=tk.LEFT, padx=(10, 0)
        )
        ttk.Checkbutton(ai_bar, text="Debug模式", variable=self.debug_mode_var, command=self._refresh_tree_columns).pack(
            side=tk.LEFT, padx=(10, 0)
        )

        row += 1
        action_bar = ttk.Frame(frm)
        action_bar.grid(row=row, column=0, columnspan=3, sticky="ew", pady=8)
        self.start_btn = ttk.Button(action_bar, text="开始处理", command=self._start_processing)
        self.start_btn.pack(side=tk.LEFT)
        ttk.Button(action_bar, text="打开表格", command=self._open_output_file).pack(side=tk.LEFT, padx=(8, 0))
        self.retry_btn = ttk.Button(action_bar, text="重试连接数据库", command=self._retry_online_connection)
        self.retry_btn.pack(side=tk.LEFT, padx=(8, 0))

        row += 1
        ttk.Label(frm, textvariable=self.progress_text_var).grid(row=row, column=0, columnspan=3, sticky="w")

        row += 1
        self.progress = ttk.Progressbar(frm, mode="determinate", maximum=100)
        self.progress.grid(row=row, column=0, columnspan=3, sticky="ew", pady=(2, 10))

        row += 1
        ttk.Label(frm, text="结果预览（前500行）").grid(row=row, column=0, columnspan=3, sticky="w", pady=(0, 4))

        row += 1
        self.all_tree_columns = ("单词", "频次", "打标", "mark", "Weight", "Total", "占比", "导入短语数据")
        self.tree = ttk.Treeview(frm, columns=self.all_tree_columns, show="headings", height=20)
        for col in self.all_tree_columns:
            self.tree.heading(col, text=col)
            if col == "mark":
                w = 300
            elif col == "导入短语数据":
                w = 360
            else:
                w = 120
            self.tree.column(col, width=w, anchor="center")
        self.tree.grid(row=row, column=0, columnspan=3, sticky="nsew")
        self._refresh_tree_columns()

        y_scroll = ttk.Scrollbar(frm, orient=tk.VERTICAL, command=self.tree.yview)
        y_scroll.grid(row=row, column=3, sticky="ns")
        self.tree.configure(yscrollcommand=y_scroll.set)

        frm.columnconfigure(1, weight=1)
        frm.rowconfigure(row, weight=1)
        self._enable_file_drag_drop(frm)

    def _enable_file_drag_drop(self, frm: ttk.Frame) -> None:
        if not DND_FILES:
            return

        def on_drop(event) -> None:
            files = _extract_dropped_files(self.root, getattr(event, "data", ""))
            if not files:
                return
            first = Path(files[0])
            if first.suffix.lower() not in {".xlsx", ".xlsm", ".csv"}:
                messagebox.showwarning("提示", "仅支持拖入 xlsx/xlsm/csv 文件。")
                return
            if not first.exists():
                messagebox.showwarning("提示", "拖入文件不存在。")
                return
            self.input_var.set(str(first))
            self._load_sheets(first)
            self.output_user_customized = False
            self.output_var.set(str(_default_output_for_input(first)))

        try:
            frm.drop_target_register(DND_FILES)
            frm.dnd_bind("<<Drop>>", on_drop)
        except Exception:
            pass

    def _choose_input(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.csv"), ("All", "*.*")])
        if not path:
            return
        self.input_var.set(path)
        self._load_sheets(Path(path))
        self.output_user_customized = False
        self.output_var.set(str(_default_output_for_input(Path(path))))

    def _choose_output(self) -> None:
        default_name = _default_output_for_input(Path(self.input_var.get()) if self.input_var.get() else None).name
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
        )
        if path:
            self.output_user_customized = True
            self.output_var.set(path)

    def _refresh_tree_columns(self) -> None:
        if bool(self.debug_mode_var.get()):
            self.tree["displaycolumns"] = self.all_tree_columns
        else:
            self.tree["displaycolumns"] = ("单词", "频次", "打标", "mark", "Weight", "Total", "占比")

    def _open_output_file(self) -> None:
        output_path = self.last_generated_output_path
        if (not output_path or not output_path.exists()) and self.output_var.get().strip():
            output_path = Path(self.output_var.get().strip())
        if not output_path or not output_path.exists():
            messagebox.showwarning("提示", "当前输出文件不存在，请先处理并生成文件。")
            return
        try:
            os.startfile(str(output_path))
        except Exception as exc:
            messagebox.showerror("错误", f"打开输出文件失败：{exc}")

    def _open_norm_manager(self) -> None:
        win = tk.Toplevel(self.root)
        win.title("归一化管理")
        win.geometry("980x620")
        win.transient(self.root)
        win.grab_set()
        notebook = ttk.Notebook(win)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        def is_plural_row(row: dict) -> bool:
            source = str(row.get("source", "")).strip().lower()
            return source.startswith("auto_discovery_plural")

        def build_tab(tab: ttk.Frame, *, plural_mode: bool) -> None:
            status_filter_var = tk.StringVar(value="all")
            keyword_var = tk.StringVar(value="")
            status_var = tk.StringVar(value="")
            rows_holder: dict[str, list[dict]] = {"rows": []}
            local_q: queue.Queue[tuple[str, dict]] = queue.Queue()

            ctrl = ttk.Frame(tab)
            ctrl.pack(fill=tk.X, pady=(0, 8))
            ttk.Label(ctrl, text="状态").pack(side=tk.LEFT)
            status_combo = ttk.Combobox(
                ctrl,
                textvariable=status_filter_var,
                state="readonly",
                values=["pending", "manual_confirmed", "auto_confirmed", "rejected", "all"],
                width=18,
            )
            status_combo.pack(side=tk.LEFT, padx=(6, 10))
            ttk.Label(ctrl, text="关键词").pack(side=tk.LEFT)
            ttk.Entry(ctrl, textvariable=keyword_var, width=30).pack(side=tk.LEFT, padx=(6, 10))

            columns = ("raw_phrase", "normalized_phrase", "status", "source", "confidence", "hit_count", "updated_at")
            tree = ttk.Treeview(tab, columns=columns, show="headings", height=20, selectmode="extended")
            headers = {
                "raw_phrase": "原短语",
                "normalized_phrase": "归一短语",
                "status": "状态",
                "source": "来源",
                "confidence": "置信度",
                "hit_count": "累计",
                "updated_at": "更新时间",
            }
            widths = {
                "raw_phrase": 170,
                "normalized_phrase": 170,
                "status": 110,
                "source": 170,
                "confidence": 90,
                "hit_count": 80,
                "updated_at": 160,
            }
            for col in columns:
                tree.heading(col, text=headers[col])
                tree.column(col, width=widths[col], anchor="center")
            tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
            y_scroll = ttk.Scrollbar(tab, orient=tk.VERTICAL, command=tree.yview)
            y_scroll.pack(fill=tk.Y, side=tk.RIGHT)
            tree.configure(yscrollcommand=y_scroll.set)

            ttk.Label(tab, textvariable=status_var).pack(anchor="w", pady=(8, 2))
            btns = ttk.Frame(tab)
            btns.pack(fill=tk.X, pady=(4, 0))

            def selected_raw_phrases() -> list[str]:
                picked = []
                for iid in tree.selection():
                    vals = tree.item(iid, "values")
                    if vals and vals[0]:
                        picked.append(str(vals[0]))
                return picked

            def redraw() -> None:
                for item in tree.get_children():
                    tree.delete(item)
                filt = keyword_var.get().strip().lower()
                current = rows_holder["rows"]
                shown = 0
                for row in current:
                    raw_phrase = str(row.get("raw_phrase", ""))
                    norm_phrase = str(row.get("normalized_phrase", ""))
                    if filt and filt not in raw_phrase.lower() and filt not in norm_phrase.lower():
                        continue
                    tree.insert(
                        "",
                        "end",
                        values=(
                            raw_phrase,
                            norm_phrase,
                            row.get("status", ""),
                            row.get("source", ""),
                            row.get("confidence", ""),
                            row.get("hit_count", ""),
                            row.get("updated_at", ""),
                        ),
                    )
                    shown += 1
                status_var.set(f"共 {len(current)} 条，当前显示 {shown} 条")

            def refresh_rows() -> None:
                status_var.set("加载中...")
                filter_value = status_filter_var.get().strip().lower()

                def worker() -> None:
                    try:
                        tagger = self._build_tagger(enable_ai=False)
                        query_status = None if filter_value == "all" else filter_value
                        rows = tagger.list_phrase_normalization(status=query_status)
                        local_q.put(("norm_rows", {"rows": rows}))
                    except Exception as exc:
                        local_q.put(("norm_error", {"error": str(exc)}))

                threading.Thread(target=worker, daemon=True).start()

                def poll() -> None:
                    try:
                        event, data = local_q.get_nowait()
                    except queue.Empty:
                        win.after(120, poll)
                        return
                    if event == "norm_error":
                        status_var.set(f"加载失败：{data.get('error', '')}")
                        messagebox.showerror("错误", str(data.get("error", "")))
                        return
                    all_rows = list(data.get("rows", []))
                    rows_holder["rows"] = [
                        x for x in all_rows if (is_plural_row(x) if plural_mode else (not is_plural_row(x)))
                    ]
                    redraw()

                win.after(120, poll)

            def apply_status(target_status: str) -> None:
                picks = selected_raw_phrases()
                if not picks:
                    messagebox.showinfo("提示", "请先在列表中选择至少一条。")
                    return
                status_var.set("提交中...")

                def worker() -> None:
                    try:
                        tagger = self._build_tagger(enable_ai=False)
                        updated = tagger.update_phrase_normalization_status(picks, target_status)
                        local_q.put(("norm_update_done", {"updated": updated}))
                    except Exception as exc:
                        local_q.put(("norm_update_error", {"error": str(exc)}))

                threading.Thread(target=worker, daemon=True).start()

                def poll() -> None:
                    try:
                        event, data = local_q.get_nowait()
                    except queue.Empty:
                        win.after(120, poll)
                        return
                    if event == "norm_update_error":
                        status_var.set(f"更新失败：{data.get('error', '')}")
                        messagebox.showerror("错误", str(data.get("error", "")))
                        return
                    status_var.set(f"更新成功：{int(data.get('updated', 0))} 条")
                    refresh_rows()

                win.after(120, poll)

            ttk.Button(btns, text="刷新", command=refresh_rows).pack(side=tk.LEFT)
            ttk.Button(btns, text="确认选中", command=lambda: apply_status("manual_confirmed")).pack(side=tk.LEFT, padx=(8, 0))
            ttk.Button(btns, text="拒绝选中", command=lambda: apply_status("rejected")).pack(side=tk.LEFT, padx=(8, 0))
            ttk.Button(btns, text="重置为待确认", command=lambda: apply_status("pending")).pack(side=tk.LEFT, padx=(8, 0))
            ttk.Button(btns, text="筛选", command=redraw).pack(side=tk.RIGHT)
            refresh_rows()

        phrase_tab = ttk.Frame(notebook, padding=10)
        plural_tab = ttk.Frame(notebook, padding=10)
        notebook.add(phrase_tab, text="词组归一化")
        notebook.add(plural_tab, text="复数归一化")
        build_tab(phrase_tab, plural_mode=False)
        build_tab(plural_tab, plural_mode=True)

    def _confirm_normalization_candidates(
        self,
        tagger: Tagger,
        candidates: list[dict],
        *,
        dialog_title: str,
        source_prefixes: tuple[str, ...],
    ) -> bool:
        allowed_sources = {str(x or "").strip().lower() for x in source_prefixes if str(x or "").strip()}
        candidate_raws = {
            str(x.get("raw_phrase", "")).strip().lower()
            for x in candidates
            if str(x.get("status", "")).lower() == "pending"
        }
        pending_seed = [x for x in candidates if str(x.get("status", "")).lower() == "pending"]
        if not pending_seed:
            return True

        win = tk.Toplevel(self.root)
        win.title(dialog_title)
        win.geometry("880x520")
        win.transient(self.root)
        win.grab_set()

        result = {"continue": False}
        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text=f"本次发现待确认 {len(pending_seed)} 条，请确认后继续处理。").pack(anchor="w", pady=(0, 8))
        checked_ids: set[str] = set()
        cols = ("checked", "raw_phrase", "normalized_phrase", "hit_count", "confidence")
        tree = ttk.Treeview(frame, columns=cols, show="headings", height=18, selectmode="extended")
        tree.heading("checked", text="☐")
        tree.heading("raw_phrase", text="原短语")
        tree.heading("normalized_phrase", text="归一短语")
        tree.heading("hit_count", text="累计")
        tree.heading("confidence", text="置信度")
        tree.column("checked", width=46, anchor="center", stretch=False)
        tree.column("raw_phrase", width=260, anchor="center")
        tree.column("normalized_phrase", width=260, anchor="center")
        tree.column("hit_count", width=100, anchor="center")
        tree.column("confidence", width=120, anchor="center")
        tree.pack(fill=tk.BOTH, expand=True)
        status_var = tk.StringVar(value="")
        continue_btn: ttk.Button | None = None

        def _set_row_checked(iid: str, checked: bool) -> None:
            if checked:
                checked_ids.add(iid)
            else:
                checked_ids.discard(iid)
            vals = list(tree.item(iid, "values"))
            if vals:
                vals[0] = "☑" if checked else "☐"
                tree.item(iid, values=vals)

        def _refresh_checked_header() -> None:
            children = tree.get_children()
            if children and all(iid in checked_ids for iid in children):
                tree.heading("checked", text="☑")
            else:
                tree.heading("checked", text="☐")

        def _refresh_status() -> None:
            pending_count = len(tree.get_children())
            selected_count = len([iid for iid in checked_ids if iid in tree.get_children()])
            if pending_count == 0:
                status_var.set("待确认 0 条，已全部完成，可继续处理")
            else:
                status_var.set(f"待确认 {pending_count} 条，已选中 {selected_count} 条")
            if continue_btn is not None:
                continue_btn.configure(state=("normal" if pending_count == 0 else "disabled"))

        def _toggle_all_checked() -> None:
            children = tree.get_children()
            if not children:
                return
            all_checked = all(iid in checked_ids for iid in children)
            for iid in children:
                _set_row_checked(iid, not all_checked)
            _refresh_checked_header()
            _refresh_status()

        tree.heading("checked", text="☐", command=_toggle_all_checked)

        def _on_tree_click(event) -> str | None:
            region = tree.identify("region", event.x, event.y)
            column = tree.identify_column(event.x)
            row_id = tree.identify_row(event.y)
            if region == "cell" and column == "#1" and row_id:
                _set_row_checked(row_id, row_id not in checked_ids)
                _refresh_checked_header()
                _refresh_status()
                return "break"
            return None

        tree.bind("<Button-1>", _on_tree_click, add="+")

        def load_pending() -> None:
            try:
                rows = tagger.list_phrase_normalization(status="pending")
            except Exception:
                rows = pending_seed
            filtered: list[dict] = []
            for row in rows:
                source = str(row.get("source", "")).strip().lower()
                raw_phrase = str(row.get("raw_phrase", "")).strip().lower()
                if candidate_raws and raw_phrase not in candidate_raws:
                    continue
                if allowed_sources and source not in allowed_sources:
                    continue
                filtered.append(row)
            checked_ids.clear()
            for item in tree.get_children():
                tree.delete(item)
            for row in filtered[:5000]:
                tree.insert(
                    "",
                    "end",
                    values=(
                        "☐",
                        row.get("raw_phrase", ""),
                        row.get("normalized_phrase", ""),
                        row.get("hit_count", ""),
                        row.get("confidence", ""),
                    ),
                )
            _refresh_checked_header()
            _refresh_status()

        load_pending()

        bar = ttk.Frame(frame)
        bar.pack(fill=tk.X, pady=(10, 0))
        ttk.Label(bar, textvariable=status_var).pack(side=tk.LEFT)

        def selected_raw_phrases() -> list[str]:
            values: list[str] = []
            for iid in tree.get_children():
                if iid not in checked_ids:
                    continue
                row = tree.item(iid, "values")
                if row and len(row) >= 2 and row[1]:
                    values.append(str(row[1]))
            return values

        def apply_selected(target_status: str) -> None:
            raw_phrases = selected_raw_phrases()
            if not raw_phrases:
                messagebox.showinfo("提示", "请先选择至少一条。")
                return
            try:
                updated = tagger.update_phrase_normalization_status(raw_phrases, target_status)
                status_var.set(f"已更新 {updated} 条")
                load_pending()
            except Exception as exc:
                messagebox.showerror("错误", f"更新归一化状态失败：{exc}")
                return

        def continue_processing() -> None:
            result["continue"] = True
            win.destroy()

        def cancel_processing() -> None:
            result["continue"] = False
            win.destroy()

        ttk.Button(bar, text="同意", command=lambda: apply_selected("manual_confirmed")).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(bar, text="拒绝", command=lambda: apply_selected("rejected")).pack(side=tk.LEFT, padx=(8, 0))
        continue_btn = ttk.Button(bar, text="继续处理", command=continue_processing, state="disabled")
        continue_btn.pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(bar, text="取消处理", command=cancel_processing).pack(side=tk.RIGHT)
        _refresh_status()

        win.protocol("WM_DELETE_WINDOW", cancel_processing)
        self.root.wait_window(win)
        return bool(result["continue"])

    def _load_sheets(self, path: Path) -> None:
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            self.sheet_combo["values"] = []
            self.sheet_var.set("")
            return
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            sheets = list(wb.sheetnames)
            self.sheet_combo["values"] = sheets
            self.sheet_var.set(sheets[0] if sheets else "")
            wb.close()
        except Exception as exc:
            messagebox.showerror("错误", f"读取Sheet失败: {exc}")

    def _open_settings_dialog(self) -> None:
        win = tk.Toplevel(self.root)
        win.title("配置")
        win.geometry("820x440")
        win.transient(self.root)
        win.grab_set()

        vars_map = {
            "openai_api_key": tk.StringVar(value=str(self.settings.get("openai_api_key", ""))),
            "ai_model": tk.StringVar(value=str(self.settings.get("ai_model", DEFAULT_AI_MODEL))),
            "ai_base_url": tk.StringVar(value=str(self.settings.get("ai_base_url", DEFAULT_AI_BASE_URL))),
            "ai_batch_size": tk.IntVar(value=int(self.settings.get("ai_batch_size", DEFAULT_BATCH_SIZE))),
            "local_cache_db": tk.StringVar(value=str(self.settings.get("local_cache_db", self.local_cache_default.name))),
            "db_host": tk.StringVar(value=str(self.settings.get("db_host", DEFAULT_DB_HOST))),
            "db_port": tk.IntVar(value=int(self.settings.get("db_port", DEFAULT_DB_PORT))),
            "db_name": tk.StringVar(value=str(self.settings.get("db_name", DEFAULT_DB_NAME))),
            "db_user": tk.StringVar(value=str(self.settings.get("db_user", DEFAULT_DB_USER))),
            "db_pass": tk.StringVar(value=str(self.settings.get("db_pass", DEFAULT_DB_PASS))),
            "db_schema": tk.StringVar(value=str(self.settings.get("db_schema", DEFAULT_DB_SCHEMA))),
            "db_table": tk.StringVar(value=str(self.settings.get("db_table", DEFAULT_DB_TABLE))),
        }
        original_api_key = self._api_key_secret or (
            "" if _is_masked_secret(str(self.settings.get("openai_api_key", ""))) else str(self.settings.get("openai_api_key", "") or "")
        )
        original_db_pass = self._db_pass_secret or (
            "" if _is_masked_secret(str(self.settings.get("db_pass", ""))) else str(self.settings.get("db_pass", "") or "")
        )

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)
        notebook = ttk.Notebook(frame)
        notebook.pack(fill=tk.BOTH, expand=True)

        ai_tab = ttk.Frame(notebook, padding=10)
        db_tab = ttk.Frame(notebook, padding=10)
        notebook.add(ai_tab, text="AI配置")
        notebook.add(db_tab, text="数据库配置")

        # AI tab
        ai_status_var = tk.StringVar(value="")
        ttk.Label(ai_tab, text="OPENAI API Key").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(ai_tab, textvariable=vars_map["openai_api_key"], show="*").grid(row=0, column=1, sticky="ew", pady=4)
        ttk.Label(ai_tab, text="AI 模型").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(ai_tab, textvariable=vars_map["ai_model"]).grid(row=1, column=1, sticky="ew", pady=4)
        ttk.Label(ai_tab, text="AI Base URL").grid(row=2, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(ai_tab, textvariable=vars_map["ai_base_url"]).grid(row=2, column=1, sticky="ew", pady=4)
        ttk.Label(ai_tab, text="批量").grid(row=3, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(ai_tab, textvariable=vars_map["ai_batch_size"]).grid(row=3, column=1, sticky="w", pady=4)
        ttk.Label(ai_tab, textvariable=ai_status_var).grid(row=4, column=0, columnspan=2, sticky="w", pady=(6, 0))

        def _resolve_secret(value: str, original_value: str) -> str:
            stripped = str(value or "").strip()
            if stripped and set(stripped) == {"*"}:
                return original_value
            return stripped

        def test_ai_connection() -> None:
            api_key = _resolve_secret(str(vars_map["openai_api_key"].get() or ""), original_api_key)
            base_url = str(vars_map["ai_base_url"].get() or "").strip().rstrip("/")
            if not api_key or not base_url:
                messagebox.showerror("错误", "请先填写 API Key 和 Base URL。")
                return
            ai_status_var.set("API测试中...")
            test_btn.configure(state="disabled")
            save_ai_btn.configure(state="disabled")
            result_holder: dict[str, str | bool] = {"done": False, "ok": False, "msg": ""}

            def worker() -> None:
                urls = [f"{base_url}/models"]
                if not base_url.endswith("/v1"):
                    urls.append(f"{base_url}/v1/models")
                headers = {"Authorization": f"Bearer {api_key}"}
                last_error = ""
                for url in urls:
                    try:
                        with httpx.Client(timeout=10.0, trust_env=False) as client:
                            resp = client.get(url, headers=headers)
                        if resp.status_code < 300:
                            result_holder["done"] = True
                            result_holder["ok"] = True
                            result_holder["msg"] = f"API连接成功：{url}"
                            return
                        if resp.status_code == 401:
                            last_error = "API Key 无效或无权限（401）"
                        else:
                            last_error = f"HTTP {resp.status_code}: {resp.text[:240]}"
                    except Exception as exc:
                        last_error = str(exc)
                result_holder["done"] = True
                result_holder["ok"] = False
                result_holder["msg"] = f"API连接失败：{last_error}"

            threading.Thread(target=worker, daemon=True).start()

            def poll_result() -> None:
                if not bool(result_holder.get("done")):
                    win.after(120, poll_result)
                    return
                test_btn.configure(state="normal")
                save_ai_btn.configure(state="normal")
                if bool(result_holder.get("ok")):
                    ai_status_var.set("API测试成功")
                    messagebox.showinfo("提示", str(result_holder.get("msg", "")))
                else:
                    ai_status_var.set("API测试失败")
                    messagebox.showerror("错误", str(result_holder.get("msg", "")))

            win.after(120, poll_result)

        def save_ai_settings() -> None:
            resolved_key = _resolve_secret(str(vars_map["openai_api_key"].get() or ""), original_api_key)
            self.settings["openai_api_key"] = resolved_key
            self._api_key_secret = resolved_key
            self.settings["ai_model"] = str(vars_map["ai_model"].get() or "").strip() or DEFAULT_AI_MODEL
            self.settings["ai_base_url"] = str(vars_map["ai_base_url"].get() or "").strip() or DEFAULT_AI_BASE_URL
            self.settings["ai_batch_size"] = max(1, int(vars_map["ai_batch_size"].get() or DEFAULT_BATCH_SIZE))
            _save_json(self.config_path, self.settings)
            ai_status_var.set("AI配置已保存")
            self.status_var.set("AI配置已保存")

        ai_btn_bar = ttk.Frame(ai_tab)
        ai_btn_bar.grid(row=5, column=0, columnspan=2, sticky="e", pady=(12, 0))
        test_btn = ttk.Button(ai_btn_bar, text="测试连接", command=test_ai_connection)
        test_btn.pack(side=tk.LEFT)
        save_ai_btn = ttk.Button(ai_btn_bar, text="保存", command=save_ai_settings)
        save_ai_btn.pack(side=tk.LEFT, padx=(8, 0))
        ai_tab.columnconfigure(1, weight=1)

        # DB tab
        db_status_var = tk.StringVar(value="")
        ttk.Label(db_tab, text="本地缓存数据库文件名").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(db_tab, textvariable=vars_map["local_cache_db"]).grid(row=0, column=1, sticky="ew", pady=4)
        ttk.Label(db_tab, text="数据库主机").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(db_tab, textvariable=vars_map["db_host"]).grid(row=1, column=1, sticky="ew", pady=4)

        line1 = ttk.Frame(db_tab)
        line1.grid(row=2, column=0, columnspan=2, sticky="ew", pady=4)
        ttk.Label(line1, text="端口").pack(side=tk.LEFT)
        ttk.Entry(line1, textvariable=vars_map["db_port"], width=8).pack(side=tk.LEFT, padx=(6, 12))
        ttk.Label(line1, text="数据库名").pack(side=tk.LEFT)
        ttk.Entry(line1, textvariable=vars_map["db_name"], width=14).pack(side=tk.LEFT, padx=(6, 12))
        ttk.Label(line1, text="用户名").pack(side=tk.LEFT)
        ttk.Entry(line1, textvariable=vars_map["db_user"], width=14).pack(side=tk.LEFT, padx=(6, 12))
        ttk.Label(line1, text="密码").pack(side=tk.LEFT)
        ttk.Entry(line1, textvariable=vars_map["db_pass"], width=14, show="*").pack(side=tk.LEFT, padx=(6, 0))

        line2 = ttk.Frame(db_tab)
        line2.grid(row=3, column=0, columnspan=2, sticky="ew", pady=4)
        ttk.Label(line2, text="Schema").pack(side=tk.LEFT)
        ttk.Entry(line2, textvariable=vars_map["db_schema"], width=14).pack(side=tk.LEFT, padx=(6, 12))
        ttk.Label(line2, text="表名").pack(side=tk.LEFT)
        ttk.Entry(line2, textvariable=vars_map["db_table"], width=30).pack(side=tk.LEFT, padx=(6, 0))

        ttk.Label(db_tab, textvariable=db_status_var).grid(row=4, column=0, columnspan=2, sticky="w", pady=(6, 0))

        def connect_and_save_db() -> None:
            host = str(vars_map["db_host"].get() or "").strip() or DEFAULT_DB_HOST
            try:
                port = int(vars_map["db_port"].get() or DEFAULT_DB_PORT)
            except Exception:
                messagebox.showerror("错误", "数据库端口格式错误。")
                return
            dbname = str(vars_map["db_name"].get() or "").strip() or DEFAULT_DB_NAME
            user = str(vars_map["db_user"].get() or "").strip() or DEFAULT_DB_USER
            password = _resolve_secret(str(vars_map["db_pass"].get() or ""), original_db_pass)
            schema = str(vars_map["db_schema"].get() or "").strip() or DEFAULT_DB_SCHEMA
            table = str(vars_map["db_table"].get() or "").strip() or DEFAULT_DB_TABLE
            local_cache_db = _normalize_cache_db_name(str(vars_map["local_cache_db"].get() or ""))

            if not _can_connect(host, port):
                db_status_var.set("连接失败")
                messagebox.showerror("错误", f"无法连接数据库 {host}:{port}，配置未保存。")
                return

            try:
                with psycopg.connect(
                    host=host,
                    user=user,
                    password=password,
                    port=port,
                    dbname=dbname,
                    connect_timeout=8,
                ) as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT 1")
                        cur.fetchone()
            except Exception as exc:
                db_status_var.set("连接失败")
                messagebox.showerror("错误", f"数据库认证/查询失败，配置未保存：\n{exc}")
                return

            self.settings["db_host"] = host
            self.settings["db_port"] = port
            self.settings["db_name"] = dbname
            self.settings["db_user"] = user
            self.settings["db_pass"] = password
            self._db_pass_secret = password
            self.settings["db_schema"] = schema
            self.settings["db_table"] = table
            self.settings["local_cache_db"] = local_cache_db
            local_db = _resolve_local_cache_db_path(local_cache_db)
            _ensure_local_cache_seed(local_db)
            _save_json(self.config_path, self.settings)
            db_status_var.set("连接成功，数据库配置已保存")
            self.status_var.set("数据库连接成功")
            self._set_mode_var("online")
            self._apply_mode_state(initial=False)
            messagebox.showinfo("提示", "数据库连接成功，已保存并切换到在线模式。")

        db_btn_bar = ttk.Frame(db_tab)
        db_btn_bar.grid(row=5, column=0, columnspan=2, sticky="e", pady=(12, 0))
        ttk.Button(db_btn_bar, text="立即连接并保存", command=connect_and_save_db).pack(side=tk.RIGHT)
        db_tab.columnconfigure(1, weight=1)

    def _apply_mode_state(self, initial: bool) -> None:
        mode = self.mode_var.get()
        self.settings["mode"] = mode
        self.settings["enable_ai"] = True
        self.settings["use_history_cache"] = bool(self.use_history_cache_var.get())
        self.settings["debug_mode"] = bool(self.debug_mode_var.get())
        self.settings["enable_phrase_normalization"] = bool(self.enable_phrase_normalization_var.get())
        self.settings["enable_plural_normalization"] = bool(self.enable_plural_normalization_var.get())
        self.settings["enable_normalization"] = bool(
            self.enable_phrase_normalization_var.get() or self.enable_plural_normalization_var.get()
        )
        _save_json(self.config_path, self.settings)

        if mode == "online":
            ok, reason = self._check_online_ready()
            self.db_available = ok
            if not ok:
                self.last_db_failure_reason = reason
                if not initial:
                    messagebox.showwarning(
                        "提示",
                        "连接数据库失败，已切换到离线模式；如果要使用在线模式，请检查数据库配置后重试。\n\n"
                        f"失败原因：{reason}",
                    )
                self._set_mode_var("offline")
                self.settings["mode"] = "offline"
                _save_json(self.config_path, self.settings)
                self.sync_done = True
                self.sync_in_progress = False
                self.status_var.set("离线模式")
                self._refresh_mode_desc()
                self._update_start_button_state()
                return
            self.last_db_failure_reason = ""
            self._refresh_mode_desc()
            self._start_online_sync_if_needed(force=not initial)
        else:
            self.sync_in_progress = False
            self.sync_done = True
            self.status_var.set("离线模式")
            self._refresh_mode_desc()
            self.progress_text_var.set("")
            self.progress["value"] = 0
            self._update_start_button_state()

    def _startup_probe_and_set_mode(self) -> None:
        ok, reason = self._check_online_ready()
        self.db_available = ok
        if ok:
            self.last_db_failure_reason = ""
            self._set_mode_var("online")
            self._apply_mode_state(initial=False)
            return
        self.last_db_failure_reason = reason
        self._set_mode_var("offline")
        self._apply_mode_state(initial=True)
        messagebox.showwarning(
            "提示",
            "连接数据库失败，已切换到离线模式；如果要使用在线模式，请检查数据库配置后重试。\n\n"
            f"失败原因：{reason}",
        )

    def _retry_online_connection(self) -> None:
        if self.is_processing:
            return
        self.status_var.set("正在重试数据库连接...")
        ok, reason = self._check_online_ready()
        self.db_available = ok
        if ok:
            self.last_db_failure_reason = ""
            self._set_mode_var("online")
            self._apply_mode_state(initial=False)
            messagebox.showinfo("提示", "数据库连接成功，已切换在线模式。")
            return
        self.last_db_failure_reason = reason
        self._set_mode_var("offline")
        self._apply_mode_state(initial=True)
        messagebox.showwarning("提示", f"数据库连接失败，当前保持离线模式。\n\n失败原因：{reason}")

    def _set_mode_var(self, mode: str) -> None:
        self._suppress_mode_event = True
        self.mode_var.set(mode)
        self._suppress_mode_event = False
        self._refresh_mode_desc()

    def _refresh_mode_desc(self) -> None:
        mode = self.mode_var.get()
        if mode == "online":
            if self.sync_in_progress:
                self.mode_desc_var.set("在线模式（同步中）")
            elif self.sync_done:
                self.mode_desc_var.set("在线模式")
            else:
                self.mode_desc_var.set("在线模式（初始化中）")
            self._update_retry_button_visibility()
            return
        if self.last_db_failure_reason:
            self.mode_desc_var.set(f"离线模式（{self.last_db_failure_reason}）")
            self._update_retry_button_visibility()
            return
        self.mode_desc_var.set("离线模式")
        self._update_retry_button_visibility()

    def _update_retry_button_visibility(self) -> None:
        if not hasattr(self, "retry_btn"):
            return
        if self.mode_var.get() == "offline":
            try:
                self.retry_btn.pack_configure(side=tk.LEFT, padx=(8, 0))
            except Exception:
                self.retry_btn.pack(side=tk.LEFT, padx=(8, 0))
        else:
            self.retry_btn.pack_forget()

    def _check_online_ready(self) -> tuple[bool, str]:
        host = str(self.settings.get("db_host", DEFAULT_DB_HOST)).strip() or DEFAULT_DB_HOST
        port = int(self.settings.get("db_port", DEFAULT_DB_PORT))
        if not _can_connect(host, port):
            return False, f"无法连接数据库 {host}:{port}"
        return True, ""

    def _start_online_sync_if_needed(self, force: bool = False) -> None:
        if not bool(self.use_history_cache_var.get()):
            self.sync_in_progress = False
            self.sync_done = True
            self.status_var.set("在线模式已就绪（不使用历史缓存）")
            self.progress_text_var.set("")
            self.progress["value"] = 0
            self._update_start_button_state()
            return
        if self.sync_in_progress:
            return
        if self.sync_done and not force:
            self._update_start_button_state()
            return

        self.sync_in_progress = True
        self.sync_done = False
        self.status_var.set("在线模式初始化中：同步数据库缓存...")
        self.progress_text_var.set("正在后台同步数据库缓存到本地，请稍候")
        self.progress["value"] = 0
        self._update_start_button_state()

        def worker() -> None:
            try:
                local_db = _resolve_local_cache_db_path(self.settings.get("local_cache_db"))
                _ensure_local_cache_seed(local_db)
                tagger = self._build_tagger(enable_ai=False)
                result = tagger.sync_all_from_db_to_local(
                    progress_callback=lambda p: self.queue.put(("sync_progress", p))
                )
                self.queue.put(("sync_done", result))
            except Exception as exc:
                self.queue.put(("sync_error", {"error": str(exc)}))

        threading.Thread(target=worker, daemon=True).start()

    def _build_tagger(self, *, enable_ai: bool) -> Tagger:
        mode = self.mode_var.get()
        local_db = _resolve_local_cache_db_path(self.settings.get("local_cache_db"))
        _ensure_local_cache_seed(local_db)

        pg = PgConfig(
            enabled=mode == "online",
            host=str(self.settings.get("db_host", DEFAULT_DB_HOST)),
            port=int(self.settings.get("db_port", DEFAULT_DB_PORT)),
            dbname=str(self.settings.get("db_name", DEFAULT_DB_NAME)),
            user=str(self.settings.get("db_user", DEFAULT_DB_USER)),
            password=str(self.settings.get("db_pass", DEFAULT_DB_PASS)),
            schema=str(self.settings.get("db_schema", DEFAULT_DB_SCHEMA)),
            table=str(self.settings.get("db_table", DEFAULT_DB_TABLE)),
        )

        return Tagger(
            enable_ai=enable_ai,
            use_history_cache=bool(self.use_history_cache_var.get()),
            model=str(self.settings.get("ai_model", DEFAULT_AI_MODEL)),
            base_url=str(self.settings.get("ai_base_url", DEFAULT_AI_BASE_URL)),
            batch_size=max(1, int(self.settings.get("ai_batch_size", DEFAULT_BATCH_SIZE))),
            local_cache_db=local_db,
            pg_config=pg,
            persist_local_on_ai=(mode == "online"),
            api_key=(self._api_key_secret or str(self.settings.get("openai_api_key", "")).strip() or None),
        )

    def _start_processing(self) -> None:
        if self.is_processing:
            return

        # Auto switch runtime mode by current DB connectivity.
        online_ok, _ = self._check_online_ready()
        expected_mode = "online" if online_ok else "offline"
        if self.mode_var.get() != expected_mode:
            self._set_mode_var(expected_mode)
            self._apply_mode_state(initial=False)

        if self.mode_var.get() == "online" and (self.sync_in_progress or not self.sync_done):
            messagebox.showinfo("提示", "在线模式缓存同步尚未完成，暂时不能开始处理。")
            return

        if not bool(self.use_history_cache_var.get()):
            api_key = str(self.settings.get("openai_api_key", "")).strip()
            if not api_key:
                messagebox.showerror("错误", "关闭历史缓存时，必须配置 API Key。")
                return

        input_path = Path(self.input_var.get().strip()) if self.input_var.get().strip() else None
        if not input_path or not input_path.exists():
            messagebox.showerror("错误", "请输入有效的输入文件路径。")
            return

        output_text = self.output_var.get().strip()
        if not self.output_user_customized:
            output_path = _default_output_for_input(input_path)
            self.output_var.set(str(output_path))
        else:
            output_path = Path(output_text) if output_text else _default_output_for_input(input_path)
        if not output_path.suffix:
            output_path = output_path.with_suffix(".xlsx")
        self.output_var.set(str(output_path))

        # Pre-stage: load rows and optional normalization confirmation.
        try:
            raw_rows = read_input_rows(input_path, self.sheet_var.get().strip() or None)
        except Exception as exc:
            messagebox.showerror("错误", f"读取输入数据失败：{exc}")
            return
        source_scope = f"{self.sheet_var.get().strip() or 'default'}::{input_path.name}"

        enable_phrase_norm = bool(self.enable_phrase_normalization_var.get())
        enable_plural_norm = bool(self.enable_plural_normalization_var.get())
        try:
            tagger_pre = self._build_tagger(enable_ai=False)
            phrase_map = tagger_pre.load_phrase_normalization_map()
            first_round_rows = normalize_input_rows_with_map(raw_rows, DEFAULT_STOPWORDS, phrase_map)

            # Round 1: phrase normalization
            if enable_phrase_norm and self.mode_var.get() == "online":
                phrase_candidates = discover_phrase_compact_normalization_candidates(
                    raw_rows,
                    DEFAULT_STOPWORDS,
                    source_scope=source_scope,
                )
                if phrase_candidates:
                    tagger_pre.record_phrase_normalization_candidates(phrase_candidates)
                    if not self._confirm_normalization_candidates(
                        tagger_pre,
                        phrase_candidates,
                        dialog_title="词组归一化确认",
                        source_prefixes=("auto_discovery",),
                    ):
                        self.status_var.set("已取消处理")
                        return
                phrase_map = tagger_pre.load_phrase_normalization_map()
                first_round_rows = normalize_input_rows_with_map(raw_rows, DEFAULT_STOPWORDS, phrase_map)

            # Round 2: plural normalization, based on round-1 output
            if enable_plural_norm and self.mode_var.get() == "online":
                plural_candidates = discover_plural_normalization_candidates(
                    first_round_rows,
                    DEFAULT_STOPWORDS,
                    source_scope=source_scope,
                )
                if plural_candidates:
                    tagger_pre.record_phrase_normalization_candidates(plural_candidates)
                    if not self._confirm_normalization_candidates(
                        tagger_pre,
                        plural_candidates,
                        dialog_title="复数归一化确认",
                        source_prefixes=("auto_discovery_plural",),
                    ):
                        self.status_var.set("已取消处理")
                        return
                phrase_map = tagger_pre.load_phrase_normalization_map()

            normalized = normalize_input_rows_with_map(raw_rows, DEFAULT_STOPWORDS, phrase_map)
        except Exception as exc:
            messagebox.showerror("错误", f"归一化预处理失败：{exc}")
            return

        self.is_processing = True
        self.status_var.set("处理中...")
        self.progress_text_var.set("准备开始")
        self.progress["value"] = 0
        self._update_start_button_state()
        self.settings["enable_ai"] = True
        self.settings["use_history_cache"] = bool(self.use_history_cache_var.get())
        self.settings["debug_mode"] = bool(self.debug_mode_var.get())
        self.settings["enable_phrase_normalization"] = bool(self.enable_phrase_normalization_var.get())
        self.settings["enable_plural_normalization"] = bool(self.enable_plural_normalization_var.get())
        self.settings["enable_normalization"] = bool(
            self.enable_phrase_normalization_var.get() or self.enable_plural_normalization_var.get()
        )
        _save_json(self.config_path, self.settings)

        def worker() -> None:
            try:
                tagger = self._build_tagger(enable_ai=True)

                def progress_cb(payload: dict) -> None:
                    self.queue.put(("process_progress", payload))

                result_rows = build_table2_with_options(
                    normalized,
                    tagger,
                    progress_callback=progress_cb,
                    source_scope=source_scope,
                    debug_mode=bool(self.debug_mode_var.get()),
                )
                write_output_with_options(
                    result_rows,
                    output_path,
                    include_debug=bool(self.debug_mode_var.get()),
                    source_raw_rows=raw_rows,
                    source_sheet_name=self.sheet_var.get().strip() or None,
                )
                self.queue.put(
                    (
                        "process_done",
                        {
                            "rows": result_rows,
                            "output": str(output_path),
                            "input_count": len(normalized),
                        },
                    )
                )
            except Exception as exc:
                self.queue.put(("process_error", {"error": str(exc)}))

        threading.Thread(target=worker, daemon=True).start()

    def _update_start_button_state(self) -> None:
        disabled = self.is_processing or (self.mode_var.get() == "online" and not self.sync_done)
        self.start_btn.configure(state=("disabled" if disabled else "normal"))

    def _poll_queue(self) -> None:
        while True:
            try:
                event, data = self.queue.get_nowait()
            except queue.Empty:
                break

            if event == "sync_progress":
                done = int(data.get("done", 0))
                total = max(1, int(data.get("total", 1)))
                pct = min(100, int(done * 100 / total))
                self.progress["value"] = pct
                self.progress_text_var.set(f"数据库同步中：{done}/{total}")

            elif event == "sync_done":
                self.sync_in_progress = False
                self.sync_done = True
                synced = int(data.get("synced_count", 0))
                synced_norm = int(data.get("synced_norm_count", 0))
                self.progress["value"] = 100
                self.progress_text_var.set(f"数据库同步完成：标签 {synced} 条，归一化 {synced_norm} 条")
                self.status_var.set("在线模式已就绪")
                self._update_start_button_state()

            elif event == "sync_error":
                self.sync_in_progress = False
                self.sync_done = False
                self._set_mode_var("offline")
                self.status_var.set("在线模式初始化失败，已切换离线模式")
                self.progress_text_var.set(data.get("error", ""))
                self._apply_mode_state(initial=False)
                messagebox.showerror("错误", f"在线模式初始化失败:\n{data.get('error', '')}")

            elif event == "process_progress":
                self._update_process_progress(data)

            elif event == "process_done":
                self.is_processing = False
                self.progress["value"] = 100
                self.status_var.set("处理完成")
                self.progress_text_var.set(f"输入{data.get('input_count', 0)}行，输出{len(data.get('rows', []))}行")
                self._update_start_button_state()
                self._refresh_preview(data.get("rows", []))
                output_done = str(data.get("output", "")).strip()
                self.last_generated_output_path = Path(output_done) if output_done else None
                if not self.output_user_customized and self.input_var.get().strip():
                    self.output_var.set(str(_default_output_for_input(Path(self.input_var.get().strip()))))
                messagebox.showinfo("完成", f"已输出: {output_done}")

            elif event == "process_error":
                self.is_processing = False
                self.status_var.set("处理失败")
                self.progress_text_var.set(data.get("error", ""))
                self.progress["value"] = 0
                self._update_start_button_state()
                messagebox.showerror("错误", data.get("error", "处理失败"))

        self.root.after(100, self._poll_queue)

    def _update_process_progress(self, payload: dict) -> None:
        stage = str(payload.get("stage", ""))
        if stage == "cache_checked":
            hit = int(payload.get("cache_hit", 0))
            missing = int(payload.get("missing", 0))
            total = int(payload.get("total_words", 0))
            self.progress["value"] = 8
            self.progress_text_var.set(f"缓存检查完成：总词数{total}，命中{hit}，待AI{missing}")
            return

        if stage == "ai_batch":
            done = int(payload.get("done_words", 0))
            total = max(1, int(payload.get("total_missing", 1)))
            pct = 10 + int(done * 80 / total)
            self.progress["value"] = min(95, pct)
            self.progress_text_var.set(
                f"AI打标中：第{payload.get('batch_index', 0)}/{payload.get('total_batches', 0)}批，{done}/{total}"
            )
            return

        if stage == "ai_done":
            self.progress["value"] = 95
            self.progress_text_var.set("AI打标完成，正在写出文件")

    def _refresh_preview(self, rows: list[dict]) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)
        preview = rows[:500]
        for row in preview:
            self.tree.insert(
                "",
                "end",
                values=(
                    row.get("单词", ""),
                    row.get("频次", ""),
                    row.get("打标", ""),
                    row.get("mark", ""),
                    row.get("Weight", ""),
                    row.get("Total", ""),
                    row.get("占比", ""),
                    row.get("导入短语数据", ""),
                ),
            )


def main() -> None:
    if TkinterDnD:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
