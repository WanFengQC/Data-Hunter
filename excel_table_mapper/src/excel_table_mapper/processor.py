from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from .tagger import Tagger, TagResult

DEFAULT_STOPWORDS = {
    "a",
    "an",
    "the",
    "and",
    "or",
    "to",
    "of",
    "for",
    "with",
    "in",
    "on",
    "at",
    "by",
    "from",
    "is",
    "are",
    "be",
    "as",
    "it",
}

TOKEN_PATTERN = re.compile(r"[a-z0-9]+")
IRREGULAR_SINGULAR_MAP = {
    "children": "child",
    "men": "man",
    "women": "woman",
    "mice": "mouse",
    "geese": "goose",
    "teeth": "tooth",
    "feet": "foot",
    "people": "person",
    "plushies": "plushie",
}


@dataclass
class RawRow:
    raw_search_term: str
    search_term: str
    weekly_exposure: float
    impressions: float
    clicks: float
    tokens: list[str]
    unique_tokens: set[str]


def _safe_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def split_words(
    text: str,
    *,
    stopwords: set[str] | None = None,
    to_lower: bool = True,
    keep_single_char: bool = False,
) -> list[str]:
    if not isinstance(text, str):
        return []
    source = text.lower() if to_lower else text
    tokens = TOKEN_PATTERN.findall(source)
    out: list[str] = []
    for token in tokens:
        if token.isdigit():
            continue
        if any(ch.isdigit() for ch in token):
            continue
        if len(token) <= 1 and not keep_single_char:
            continue
        if stopwords and token in stopwords:
            continue
        out.append(token)
    return out


def normalize_token_form(token: str) -> str:
    word = str(token or "").strip().lower()
    if not word:
        return ""
    if word in IRREGULAR_SINGULAR_MAP:
        return IRREGULAR_SINGULAR_MAP[word]
    if len(word) <= 3:
        return word
    if word.endswith("ies") and len(word) > 4:
        return f"{word[:-3]}y"
    if word.endswith("es") and len(word) > 4:
        stem = word[:-2]
        if stem.endswith(("s", "x", "z", "ch", "sh", "o")):
            return stem
    if word.endswith("s") and len(word) > 3 and not word.endswith(("ss", "us", "is")):
        return word[:-1]
    return word


def read_input_rows(input_path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    if input_path.suffix.lower() == ".csv":
        with input_path.open("r", encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))

    wb = load_workbook(input_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows, [])]
    result: list[dict[str, Any]] = []
    for row in rows:
        record = {}
        for i, key in enumerate(headers):
            if not key:
                continue
            record[key] = row[i] if i < len(row) else None
        if record:
            result.append(record)
    return result


def normalize_input_rows(raw_rows: list[dict[str, Any]], stopwords: set[str]) -> list[RawRow]:
    return normalize_input_rows_with_map(raw_rows, stopwords, {})


def _normalize_phrase_text(value: str) -> str:
    tokens = split_words(value, stopwords=None, to_lower=True, keep_single_char=True)
    return " ".join(tokens)


def _apply_phrase_map(tokens: list[str], normalization_map: dict[str, str]) -> list[str]:
    if not tokens or not normalization_map:
        return list(tokens)

    mapping_tokens: list[tuple[list[str], list[str]]] = []
    for raw_phrase, normalized_phrase in normalization_map.items():
        raw_tokens = split_words(str(raw_phrase or ""), stopwords=None, to_lower=True, keep_single_char=True)
        norm_tokens = split_words(str(normalized_phrase or ""), stopwords=None, to_lower=True, keep_single_char=True)
        if not raw_tokens or not norm_tokens:
            continue
        if raw_tokens == norm_tokens:
            continue
        mapping_tokens.append((raw_tokens, norm_tokens))
    if not mapping_tokens:
        return list(tokens)

    # Longest-phrase-first replacement on token stream.
    # Run multiple rounds so chain rules can converge:
    # e.g. "jelly cats" -> "jelly cat" -> "jellycat".
    mapping_tokens.sort(key=lambda x: len(x[0]), reverse=True)
    current = list(tokens)
    max_rounds = 4
    for _ in range(max_rounds):
        out: list[str] = []
        idx = 0
        n = len(current)
        changed = False
        while idx < n:
            replaced = False
            for raw_tokens, norm_tokens in mapping_tokens:
                m = len(raw_tokens)
                if m == 0 or idx + m > n:
                    continue
                if current[idx : idx + m] == raw_tokens:
                    out.extend(norm_tokens)
                    idx += m
                    replaced = True
                    changed = True
                    break
            if not replaced:
                out.append(current[idx])
                idx += 1
        current = out
        if not changed:
            break
    return current


def normalize_input_rows_with_map(
    raw_rows: list[dict[str, Any]],
    stopwords: set[str],
    normalization_map: dict[str, str] | None = None,
) -> list[RawRow]:
    required_aliases = {
        "search_term": ("搜索词", "keyword", "关键词", "search_term"),
        "weekly_exposure": ("预估周曝光量", "weekly_exposure"),
        "impressions": ("展示量", "impressions"),
        "clicks": ("点击量", "clicks"),
    }

    def pick(row: dict[str, Any], aliases: tuple[str, ...]) -> Any:
        for alias in aliases:
            if alias in row:
                return row.get(alias)
        return None

    normalized: list[RawRow] = []
    norm_map = normalization_map or {}
    for row in raw_rows:
        search_term = str(pick(row, required_aliases["search_term"]) or "").strip()
        if not search_term:
            continue
        source_tokens = split_words(search_term, stopwords=None, to_lower=True, keep_single_char=True)
        if not source_tokens:
            continue
        mapped_tokens = _apply_phrase_map(source_tokens, norm_map)
        if not mapped_tokens:
            continue
        normalized_term = " ".join(mapped_tokens)
        tokens = [t for t in mapped_tokens if (len(t) > 1 and (not stopwords or t not in stopwords))]
        if not tokens:
            continue
        normalized.append(
            RawRow(
                raw_search_term=search_term,
                search_term=normalized_term,
                weekly_exposure=_safe_float(pick(row, required_aliases["weekly_exposure"])),
                impressions=_safe_float(pick(row, required_aliases["impressions"])),
                clicks=_safe_float(pick(row, required_aliases["clicks"])),
                tokens=tokens,
                unique_tokens=set(tokens),
            )
        )

    if not normalized:
        raise RuntimeError("输入数据为空，或缺少可解析的“搜索词”列。")
    return normalized


def discover_phrase_compact_normalization_candidates(
    raw_rows: list[dict[str, Any]],
    stopwords: set[str],
    source_scope: str | None = None,
) -> list[dict[str, Any]]:
    required_aliases = {
        "search_term": ("搜索词", "keyword", "关键词", "search_term"),
        "weekly_exposure": ("预估周曝光量", "weekly_exposure"),
    }

    def pick(row: dict[str, Any], aliases: tuple[str, ...]) -> Any:
        for alias in aliases:
            if alias in row:
                return row.get(alias)
        return None

    # compact_key -> phrase -> stats
    grouped: dict[str, dict[str, dict[str, float]]] = defaultdict(dict)
    for row in raw_rows:
        search_term = str(pick(row, required_aliases["search_term"]) or "").strip()
        if not search_term:
            continue
        tokens = split_words(search_term, stopwords=None, to_lower=True, keep_single_char=True)
        if not tokens:
            continue
        exposure_val = _safe_float(pick(row, required_aliases["weekly_exposure"]))
        phrase = " ".join(tokens)
        compact = "".join(tokens)
        if len(compact) < 4:
            continue
        if phrase not in grouped[compact]:
            grouped[compact][phrase] = {"count": 0.0, "exposure": 0.0}
        grouped[compact][phrase]["count"] += 1.0
        grouped[compact][phrase]["exposure"] += exposure_val

    out: list[dict[str, Any]] = []
    normalized_source_scope = str(source_scope or "").strip()
    for variants in grouped.values():
        if len(variants) < 2:
            continue
        ranked = sorted(
            variants.items(),
            key=lambda kv: ((1 if " " not in kv[0] else 0), kv[1]["exposure"], kv[1]["count"]),
            reverse=True,
        )
        canonical_phrase, canonical_stats = ranked[0]
        canonical_count = int(canonical_stats["count"])
        for phrase, stats in ranked[1:]:
            if phrase == canonical_phrase:
                continue
            count = int(stats["count"])
            exposure = float(stats["exposure"])
            confidence = min(
                1.0,
                0.35
                + (0.25 if " " not in canonical_phrase else 0.0)
                + 0.1 * min(3, canonical_count)
                + 0.1 * min(3, count),
            )
            auto_confirm = (" " in phrase) and (" " not in canonical_phrase) and canonical_count >= 2 and count >= 2
            status = "auto_confirmed" if auto_confirm else "pending"
            out.append(
                {
                    "raw_phrase": phrase,
                    "normalized_phrase": canonical_phrase,
                    "status": status,
                    "source": "auto_discovery",
                    "source_scope": normalized_source_scope or None,
                    "confidence": round(confidence, 4),
                    "hit_count": max(1, count),
                    "weekly_exposure": exposure,
                }
            )
    return out


def discover_plural_normalization_candidates(
    rows: list[RawRow],
    stopwords: set[str],
    source_scope: str | None = None,
) -> list[dict[str, Any]]:
    # singular_key -> variant -> stats (for cat/cats style manual normalization candidates)
    singular_grouped: dict[str, dict[str, dict[str, float]]] = defaultdict(dict)
    for row in rows:
        unique_tokens = set(row.unique_tokens)
        for token in unique_tokens:
            singular = normalize_token_form(token)
            if not singular:
                continue
            if stopwords and (token in stopwords or singular in stopwords):
                continue
            if token not in singular_grouped[singular]:
                singular_grouped[singular][token] = {"count": 0.0, "exposure": 0.0}
            singular_grouped[singular][token]["count"] += 1.0
            singular_grouped[singular][token]["exposure"] += row.weekly_exposure

    out: list[dict[str, Any]] = []
    normalized_source_scope = str(source_scope or "").strip()
    for variants in singular_grouped.values():
        if len(variants) < 2:
            continue
        # Keep the higher-frequency form as canonical (single/plural is not forced).
        # Exposure is only used as a tie-breaker.
        ranked = sorted(
            variants.items(),
            key=lambda kv: (kv[1]["count"], kv[1]["exposure"], -len(kv[0])),
            reverse=True,
        )
        canonical_phrase, canonical_stats = ranked[0]
        canonical_count = int(canonical_stats["count"])
        for phrase, stats in ranked[1:]:
            if phrase == canonical_phrase:
                continue
            count = int(stats["count"])
            exposure = float(stats["exposure"])
            confidence = min(
                1.0,
                0.45
                + 0.1 * min(3, canonical_count)
                + 0.1 * min(3, count),
            )
            out.append(
                {
                    "raw_phrase": phrase,
                    "normalized_phrase": canonical_phrase,
                    "status": "pending",
                    "source": "auto_discovery_plural",
                    "source_scope": normalized_source_scope or None,
                    "confidence": round(confidence, 4),
                    "hit_count": max(1, count),
                    "weekly_exposure": exposure,
                }
            )
    return out


def discover_phrase_normalization_candidates(
    raw_rows: list[dict[str, Any]],
    stopwords: set[str],
    source_scope: str | None = None,
) -> list[dict[str, Any]]:
    # Backward-compatible combined discovery.
    phrase_items = discover_phrase_compact_normalization_candidates(raw_rows, stopwords, source_scope)
    first_pass_rows = normalize_input_rows_with_map(raw_rows, stopwords, {})
    plural_items = discover_plural_normalization_candidates(first_pass_rows, stopwords, source_scope)
    return [*phrase_items, *plural_items]


def build_table2(rows: list[RawRow], tagger: Tagger, progress_callback=None) -> list[dict[str, Any]]:
    return build_table2_with_options(rows, tagger, progress_callback=progress_callback, source_scope=None, debug_mode=False)


def build_table2_with_options(
    rows: list[RawRow],
    tagger: Tagger,
    *,
    progress_callback=None,
    source_scope: str | None = None,
    debug_mode: bool = False,
) -> list[dict[str, Any]]:
    freq_counter = Counter()
    for row in rows:
        freq_counter.update(row.tokens)

    word_total_clicks = defaultdict(float)
    word_total_impressions = defaultdict(float)
    word_total_weekly_exposure = defaultdict(float)

    for row in rows:
        for word in row.unique_tokens:
            word_total_clicks[word] += row.clicks
            word_total_impressions[word] += row.impressions
            word_total_weekly_exposure[word] += row.weekly_exposure

    words = [w for w, _ in freq_counter.most_common()]

    # Build per-word phrase hints from current input sheet only:
    # exact token match (via split_words) and top10 by weekly_exposure.
    phrase_candidates: dict[str, list[tuple[float, str]]] = defaultdict(list)
    for row in rows:
        for word in row.unique_tokens:
            phrase_candidates[word].append((row.weekly_exposure, row.search_term))
    phrase_hints: dict[str, list[str]] = {}
    for word, items in phrase_candidates.items():
        items_sorted = sorted(items, key=lambda x: x[0], reverse=True)
        phrase_hints[word] = [text for _, text in items_sorted[:10]]

    tags: dict[str, TagResult] = tagger.classify(
        words,
        phrase_hints=phrase_hints,
        progress_callback=progress_callback,
        source_scope=source_scope,
    )

    output: list[dict[str, Any]] = []
    for word in words:
        total = word_total_clicks[word]
        denominator = word_total_impressions[word]
        exposure_sum = word_total_weekly_exposure[word]

        weight: float | str = ""
        ratio: float | str = ""
        if denominator > 0 and total > 0:
            weight = round((total / denominator) * (4.3 * exposure_sum) * 10, 0)
            if total:
                ratio = weight / total

        tag = tags.get(word) or TagResult(tag_label="解析失败", reason=f"“{word}”，无缓存且未启用AI。")
        output.append(
            {
                "单词": word,
                "频次": int(freq_counter[word]),
                "打标": tag.tag_label,
                "mark": tag.reason,
                "导入短语数据": " | ".join(phrase_hints.get(word, [])) if debug_mode else "",
                "Weight": weight,
                "Total": round(total, 6),
                "占比": ratio,
            }
        )

    output.sort(key=lambda x: (x["频次"], float(x["Total"] or 0)), reverse=True)
    return output


def write_output(
    rows: list[dict[str, Any]],
    output_path: Path,
    *,
    source_raw_rows: list[dict[str, Any]] | None = None,
    source_sheet_name: str | None = None,
) -> None:
    return write_output_with_options(
        rows,
        output_path,
        include_debug=False,
        source_raw_rows=source_raw_rows,
        source_sheet_name=source_sheet_name,
    )


def write_output_with_options(
    rows: list[dict[str, Any]],
    output_path: Path,
    *,
    include_debug: bool = False,
    source_raw_rows: list[dict[str, Any]] | None = None,
    source_sheet_name: str | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    columns = ["单词", "频次", "打标", "mark", "Weight", "Total", "占比"]
    if include_debug:
        columns.append("导入短语数据")
    if output_path.suffix.lower() == ".csv":
        with output_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            writer.writerows(rows)
        return

    wb = Workbook()
    ws_source = wb.active
    ws_source.title = "sheet1"
    src_rows = source_raw_rows or []
    if src_rows:
        header_order: list[str] = []
        seen = set()
        for r in src_rows:
            for k in r.keys():
                key = str(k or "").strip()
                if not key or key in seen:
                    continue
                seen.add(key)
                header_order.append(key)
        if header_order:
            ws_source.append(header_order)
            for r in src_rows:
                ws_source.append([r.get(h, "") for h in header_order])

    ws = wb.create_sheet("sheet2")

    # Row 1 blank, row 2 summary, row 3 headers, row 4+ data.
    ws.append([])
    weight_sum = sum(float(r.get("Weight") or 0) for r in rows if isinstance(r.get("Weight"), (int, float)))
    total_sum = sum(float(r.get("Total") or 0) for r in rows if isinstance(r.get("Total"), (int, float)))
    ratio_sum = (weight_sum / total_sum) if total_sum else ""
    ws.append(["", "", "", "", weight_sum if weight_sum else "", total_sum if total_sum else "", ratio_sum])
    ws["G2"].number_format = "0%"

    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])

    last_row = ws.max_row
    for idx in range(4, last_row + 1):
        cell_val = ws[f"G{idx}"].value
        if isinstance(cell_val, (int, float)):
            ws[f"G{idx}"].number_format = "0%"
    if last_row >= 4:
        ws.conditional_formatting.add(
            f"G4:G{last_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="00B050",
                mid_type="num",
                mid_value=0.5,
                mid_color="FFD966",
                end_type="num",
                end_value=1.5,
                end_color="FF0000",
            ),
        )

    # Auto-fit column widths based on rendered text length.
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws[f"{col_letter}{row_idx}"].value
            text = "" if value is None else str(value)
            max_len = max(max_len, len(text))
        ws.column_dimensions[col_letter].width = min(80, max(8, max_len + 2))

    wb.save(output_path)

