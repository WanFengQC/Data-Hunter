from __future__ import annotations

import argparse
import csv
import json
import os
import re
import time
from collections import deque
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import psycopg
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook, load_workbook


WORD_TOKEN = re.compile(r"[a-z0-9]+")
VALID_TAGS = {"1核心词", "2外形", "3属性", "4痛点", "5规格", "6受众", "7场景", "8品牌", "无效词"}


@dataclass
class RowData:
    raw: dict[str, Any]
    word: str


def norm_word(text: str) -> str:
    return str(text or "").strip().lower()


def clean_phrase(text: str) -> str:
    value = str(text or "")
    value = value.replace(chr(65532), "")
    return value.strip().strip('"').strip().lower()


def parse_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").strip('"')
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def load_rows(input_path: Path, sheet_name: str | None = None) -> tuple[list[RowData], list[str]]:
    if input_path.suffix.lower() == ".csv":
        with input_path.open("r", encoding="utf-8-sig", newline="") as f:
            raw_rows = list(csv.DictReader(f))
        headers = list(raw_rows[0].keys()) if raw_rows else []
        rows: list[RowData] = []
        for row in raw_rows:
            word = norm_word(row.get("单词") or row.get("word") or row.get("关键词") or "")
            if not word:
                continue
            rows.append(RowData(raw=row, word=word))
        return rows, headers

    wb = load_workbook(input_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header_row_idx = None
    headers: list[str] = []
    for r in range(1, min(ws.max_row, 20) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        row_headers = [str(v).strip() if v is not None else "" for v in row_vals]
        if "单词" in row_headers or "word" in row_headers or "关键词" in row_headers:
            header_row_idx = r
            headers = row_headers
            break
    if header_row_idx is None:
        raise RuntimeError("未找到表头（需要包含“单词/word/关键词”列）。")

    rows: list[RowData] = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        raw: dict[str, Any] = {}
        has_value = False
        for c, key in enumerate(headers, start=1):
            if not key:
                continue
            val = ws.cell(row=r, column=c).value
            raw[key] = val
            if val not in (None, ""):
                has_value = True
        if not has_value:
            continue
        word = norm_word(raw.get("单词") or raw.get("word") or raw.get("关键词") or "")
        if not word:
            continue
        rows.append(RowData(raw=raw, word=word))
    return rows, headers


def chunked(items: list[str], size: int) -> list[list[str]]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def build_word_boundary_regex(words: list[str]) -> str:
    safe_words = [re.escape(w) for w in words if w]
    # Match whole token based on non-alnum boundaries.
    return rf"(^|[^a-z0-9])(?:{'|'.join(safe_words)})([^a-z0-9]|$)"


def fetch_top_phrase_refs(
    *,
    words: list[str],
    host: str,
    port: int,
    dbname: str,
    user: str,
    password: str,
    schema: str,
    table: str,
    max_refs_per_word: int,
    regex_chunk_size: int,
    max_rows_per_chunk: int,
) -> dict[str, list[tuple[str, float]]]:
    if not words:
        return {}
    uniq_words = sorted(set(words))
    refs: dict[str, list[tuple[str, float]]] = {w: [] for w in uniq_words}
    phrase_best_searches: dict[str, float] = {}

    keyword_expr = "lower(trim(both '\"' from replace(CAST(keyword AS TEXT), chr(65532), '')))"
    searches_expr = "NULLIF(trim(both '\"' from replace(CAST(searches AS TEXT), chr(65532), '')), '')::double precision"

    with psycopg.connect(host=host, port=port, dbname=dbname, user=user, password=password, connect_timeout=30) as conn:
        with conn.cursor() as cur:
            for group in chunked(uniq_words, regex_chunk_size):
                regex = build_word_boundary_regex(group)
                query = f"""
                    SELECT {keyword_expr} AS phrase, {searches_expr} AS searches
                    FROM {schema}.{table}
                    WHERE keyword IS NOT NULL
                      AND searches IS NOT NULL
                      AND {keyword_expr} ~ %s
                    ORDER BY {searches_expr} DESC NULLS LAST
                    LIMIT %s
                """
                cur.execute(query, (regex, max_rows_per_chunk))
                rows = cur.fetchall()
                for phrase_raw, searches_raw in rows:
                    phrase = clean_phrase(phrase_raw)
                    searches = parse_float(searches_raw)
                    if not phrase or searches is None:
                        continue
                    prev = phrase_best_searches.get(phrase)
                    if prev is None or searches > prev:
                        phrase_best_searches[phrase] = searches

    sorted_phrases = sorted(phrase_best_searches.items(), key=lambda x: (x[1], x[0]), reverse=True)
    for phrase, searches in sorted_phrases:
        tokens = set(WORD_TOKEN.findall(phrase))
        if not tokens:
            continue
        matched_words = tokens.intersection(uniq_words)
        if not matched_words:
            continue
        for word in matched_words:
            if len(refs[word]) >= max_refs_per_word:
                continue
            if phrase == word:
                continue
            refs[word].append((phrase, searches))
    return refs


def build_prompt(batch: list[str], refs_map: dict[str, list[tuple[str, float]]]) -> str:
    lines: list[str] = []
    for word in batch:
        refs = refs_map.get(word, [])
        if refs:
            refs_text = "; ".join(f"{p} ({int(s)})" for p, s in refs)
        else:
            refs_text = "无"
        lines.append(f"关键词: {word} | 参考短语(按搜索量降序): {refs_text}")

    payload = "\n\n".join(lines)
    return f"""
# 亚马逊词根打标

请将下面的英文词根映射为唯一标签。站在10年跨境PM视角，为每个标签简述为何选择该标签。

标签必须是以下之一：
1核心词 / 2外形 / 3属性 / 4痛点 / 5规格 / 6受众 / 7场景 / 8品牌 / 无效词

标签定义：
- 1核心词：产品的核心名称或类别名（如：stuffed, plush, animal, toy, pillow）。
- 2外形：视觉主体形象、物种名（如：bear, panda, cat, unicorn, dinosaur）。
- 3属性：产品的物理特性、改性工艺或硬件卖点（如：weighted, heated, warmable, cooling）。
- 4痛点：用户主观渴望解决的情绪、生理问题或购买诉求（如：anxiety, sleep, sensory, stress）。
- 5规格：描述产品的大小、重量或厚度规格（如：giant, large, 5lbs, mini）。
- 6受众：描述产品的使用人群（如：adults, kids, toddler, baby）。
- 7场景：描述产品的使用时机或赠送节点（如：gift, birthday, valentine）。
- 8品牌：描述特定的品牌名称或受保护的IP（如：warmies, disney, squishmallow）。
- 无效词：语义不完整或业务价值低，按无效词处理。

返回要求：
1. 只返回 JSON。
2. 顺序必须和输入一致。
3. 原因备注用简短中文说明，不要输出多余解释。
4. 每行有“关键词 + 参考短语(Top10)”；参考短语仅作为该关键词语义判定依据，不改变标签集合。

格式：
{{
  "rows": [
    {{
      "关键词": "stuffed",
      "对应标签": "1核心词",
      "原因备注": "“填充”，类目核心词。"
    }},
    {{
      "关键词": "weighted",
      "对应标签": "3属性",
      "原因备注": "“加重”，产品核心属性。"
    }},
    {{
      "关键词": "pink",
      "对应标签": "5规格",
      "原因备注": "“粉色”，颜色规格词，用于区分配色款式。"
    }}
  ]
}}

词根列表：
{payload}
""".strip()


def parse_ai_rows(content: str) -> dict[str, dict[str, str]]:
    text = str(content or "").strip()
    if text.startswith("```"):
        lines = text.splitlines()
        if len(lines) >= 3:
            text = "\n".join(lines[1:-1]).strip()
    try:
        payload = json.loads(text)
    except Exception:
        return {}
    rows = payload.get("rows") if isinstance(payload, dict) else None
    if not isinstance(rows, list):
        return {}
    out: dict[str, dict[str, str]] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        word = norm_word(row.get("关键词") or row.get("keyword") or row.get("word") or "")
        if not word:
            continue
        out[word] = {
            "translation_zh": str(row.get("translation_zh") or row.get("翻译") or "").strip(),
            "tag_label": str(row.get("对应标签") or row.get("标签") or row.get("label") or "").strip(),
            "reason": str(row.get("原因备注") or row.get("原因") or row.get("reason") or "").strip(),
        }
    return out


def ai_relabel(
    *,
    words: list[str],
    refs_map: dict[str, list[tuple[str, float]]],
    api_key: str,
    base_url: str,
    model: str,
    batch_size: int,
    max_word_retries: int,
) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, Any]]]:
    client = OpenAI(api_key=api_key, base_url=base_url)
    output: dict[str, dict[str, str]] = {}
    failed: dict[str, dict[str, Any]] = {}
    pending = deque(words)
    attempts: dict[str, int] = {w: 0 for w in words}

    def _is_quota_error(err: Exception | None) -> bool:
        if err is None:
            return False
        text = str(err).lower()
        return ("insufficient_quota" in text) or ("quota" in text) or ("额度" in text) or ("billing" in text)

    def _is_valid_row(row: dict[str, str] | None) -> bool:
        if not isinstance(row, dict):
            return False
        tag_label = str(row.get("tag_label") or "").strip()
        reason = str(row.get("reason") or "").strip()
        if tag_label not in VALID_TAGS:
            return False
        if not reason:
            return False
        return True

    while pending:
        batch: list[str] = []
        while pending and len(batch) < max(1, batch_size):
            word = pending.popleft()
            if word in output or word in failed:
                continue
            attempts[word] = int(attempts.get(word, 0)) + 1
            batch.append(word)
        if not batch:
            break

        prompt = build_prompt(batch, refs_map)
        resp = None
        last_err: Exception | None = None
        quota_exhausted = False
        for attempt in range(1, 5):
            try:
                resp = client.chat.completions.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": "你是严格的电商词根打标助手，只输出JSON。"},
                        {"role": "user", "content": prompt},
                    ],
                    temperature=0,
                )
                last_err = None
                break
            except Exception as ex:
                last_err = ex
                if _is_quota_error(ex):
                    quota_exhausted = True
                    break
                if attempt >= 4:
                    break
                time.sleep(min(12, 2 * attempt))

        if resp is None:
            if quota_exhausted:
                for word in batch:
                    failed[word] = {
                        "attempts": attempts.get(word, 0),
                        "reason": "API额度不足或计费受限，未完成打标。",
                    }
                while pending:
                    rest = pending.popleft()
                    if rest in output or rest in failed:
                        continue
                    failed[rest] = {
                        "attempts": attempts.get(rest, 0),
                        "reason": "API额度不足或计费受限，未完成打标。",
                    }
                break

            for word in batch:
                if attempts.get(word, 0) >= max_word_retries:
                    failed[word] = {
                        "attempts": attempts.get(word, 0),
                        "reason": f"AI请求失败: {last_err}",
                    }
                else:
                    pending.append(word)
            continue

        content = (resp.choices[0].message.content or "").strip()
        parsed = parse_ai_rows(content)
        for word in batch:
            row = parsed.get(word)
            if _is_valid_row(row):
                output[word] = row
                continue
            if attempts.get(word, 0) >= max_word_retries:
                failed[word] = {
                    "attempts": attempts.get(word, 0),
                    "reason": f"AI返回无效或缺失结果: {word}",
                }
            else:
                pending.append(word)

    return output, failed


def write_output_xlsx(
    *,
    rows: list[RowData],
    headers: list[str],
    refs_map: dict[str, list[tuple[str, float]]],
    relabel_map: dict[str, dict[str, str]],
    failed_map: dict[str, dict[str, Any]],
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    base_headers = [h for h in headers if h]
    if not base_headers:
        base_headers = ["单词"]

    add_headers = [
        "二次校对_翻译(Toys&Games)",
        "二次校对_标签",
        "二次校对_原因",
        "二次校对_参考短语Top10",
    ]
    final_headers = base_headers + add_headers

    wb = Workbook()
    ws = wb.active
    ws.title = "二次校对"
    ws.append(final_headers)

    for item in rows:
        if item.word in failed_map:
            continue
        base_vals = [item.raw.get(h, "") for h in base_headers]
        refs = refs_map.get(item.word, [])
        refs_text = " | ".join([f"{p} ({int(s)})" for p, s in refs]) if refs else ""
        relabel = relabel_map.get(item.word, {})
        ws.append(
            base_vals
            + [
                relabel.get("translation_zh", ""),
                relabel.get("tag_label", ""),
                relabel.get("reason", ""),
                refs_text,
            ]
        )

    fail_ws = wb.create_sheet("失败词")
    fail_ws.append(["单词", "尝试次数", "失败原因", "原始打标", "参考短语Top10"])
    row_map: dict[str, RowData] = {r.word: r for r in rows}
    for word in sorted(failed_map.keys()):
        item = row_map.get(word)
        refs = refs_map.get(word, [])
        refs_text = " | ".join([f"{p} ({int(s)})" for p, s in refs]) if refs else ""
        old_tag = ""
        if item:
            old_tag = str(item.raw.get("打标") or item.raw.get("tag_label") or "")
        fail_info = failed_map.get(word) or {}
        fail_ws.append(
            [
                word,
                int(fail_info.get("attempts") or 0),
                str(fail_info.get("reason") or ""),
                old_tag,
                refs_text,
            ]
        )
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="基于 ABA 短语参考做词根二次校对打标，不覆盖旧数据。")
    parser.add_argument("--input", required=True, help="输入文件（xlsx/csv）")
    parser.add_argument("--output", required=True, help="输出文件（xlsx）")
    parser.add_argument("--sheet", default="", help="输入sheet名（xlsx可选）")
    parser.add_argument("--pg-host", default=os.getenv("PG_HOST", "192.168.110.107"))
    parser.add_argument("--pg-port", type=int, default=int(os.getenv("PG_PORT", "5432")))
    parser.add_argument("--pg-db", default=os.getenv("PG_DB", "hunter"))
    parser.add_argument("--pg-user", default=os.getenv("PG_USER", "postgres"))
    parser.add_argument("--pg-pass", default=os.getenv("PG_PASS", "123456"))
    parser.add_argument("--pg-schema", default=os.getenv("PG_SCHEMA", "public"))
    parser.add_argument("--pg-table", default=os.getenv("PG_TABLE", "seller_sprite_items"))
    parser.add_argument("--max-refs-per-word", type=int, default=10)
    parser.add_argument("--regex-chunk-size", type=int, default=80)
    parser.add_argument("--max-rows-per-chunk", type=int, default=200000)
    parser.add_argument("--batch-size", type=int, default=80)
    parser.add_argument("--max-word-retries", type=int, default=3, help="每个词最多尝试次数（含首次）")
    parser.add_argument("--model", default=os.getenv("AI_MODEL", "gpt-5.4"))
    parser.add_argument("--base-url", default=os.getenv("AI_BASE_URL", "https://api.wfqc8.cn/v1"))
    parser.add_argument("--api-key", default=os.getenv("OPENAI_API_KEY", ""))
    parser.add_argument("--skip-ai", action="store_true", help="仅生成短语参考，不调用AI。")
    return parser.parse_args()


def main() -> None:
    load_dotenv()
    args = parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    rows, headers = load_rows(input_path, args.sheet.strip() or None)
    words = sorted({r.word for r in rows})
    if not words:
        raise RuntimeError("未读取到可处理词语。")

    refs_map = fetch_top_phrase_refs(
        words=words,
        host=args.pg_host,
        port=int(args.pg_port),
        dbname=args.pg_db,
        user=args.pg_user,
        password=args.pg_pass,
        schema=args.pg_schema,
        table=args.pg_table,
        max_refs_per_word=max(1, int(args.max_refs_per_word)),
        regex_chunk_size=max(10, int(args.regex_chunk_size)),
        max_rows_per_chunk=max(10000, int(args.max_rows_per_chunk)),
    )

    relabel_map: dict[str, dict[str, str]] = {}
    failed_map: dict[str, dict[str, Any]] = {}
    if not args.skip_ai:
        api_key = str(args.api_key or "").strip()
        if not api_key:
            raise RuntimeError("未检测到 OPENAI_API_KEY，请设置后重试，或加 --skip-ai 先生成参考短语。")
        relabel_map, failed_map = ai_relabel(
            words=words,
            refs_map=refs_map,
            api_key=api_key,
            base_url=args.base_url,
            model=args.model,
            batch_size=max(1, int(args.batch_size)),
            max_word_retries=max(1, int(args.max_word_retries)),
        )

    write_output_xlsx(
        rows=rows,
        headers=headers,
        refs_map=refs_map,
        relabel_map=relabel_map,
        failed_map=failed_map,
        output_path=output_path,
    )

    words_with_refs = sum(1 for w in words if refs_map.get(w))
    print(f"输入词数: {len(words)}")
    print(f"有参考短语的词数: {words_with_refs}")
    if args.skip_ai:
        print("AI打标: 已跳过（--skip-ai）")
    else:
        print(f"AI打标完成词数: {len(relabel_map)}")
        print(f"失败词数: {len(failed_map)}")
    print(f"输出文件: {output_path}")


if __name__ == "__main__":
    main()
